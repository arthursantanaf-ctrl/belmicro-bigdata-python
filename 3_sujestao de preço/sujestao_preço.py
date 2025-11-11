import pandas as pd
import numpy as np
import os
import re

# Importar bibliotecas do openpyxl para formatação
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

# --- 1. CONFIGURAÇÃO ---
ARQUIVO_ENTRADA = r"C:\Users\asf\Documents\resultado final shopee\limpeza coleta\resultados_shopee_finalissimo.xlsx"
ARQUIVO_SAIDA = r"C:\Users\asf\Documents\resultado final shopee\limpeza coleta\REsLATORIO_CORRIGIDO_V5_FINAL.xlsx" # Novo nome de saída (V5)

# Coluna que tem a análise "SIM" / "NÃO"
COLUNA_ANALISE = "Comparativo" 

print("🏁 INICIANDO SCRIPT DE SUGESTÃO DE PREÇO (V5 - Incluindo Belmicro) 🏁")

# --- 2. FUNÇÃO DE LIMPEZA DE PREÇO ---
def limpar_preco(valor):
    """Converte texto/número de preço em float, corrigindo formatos e erros de 100x."""
    if pd.isna(valor): return np.nan
    if isinstance(valor, (int, float)):
        if valor > 50000: return valor / 100.0
        return float(valor)
    s = str(valor).replace("R$", "").strip()
    if ',' in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    try:
        preco = float(s)
        if preco > 50000: return preco / 100.0
        return preco
    except Exception: return np.nan

# --- 3. LER PLANILHA E PREPARAR DADOS ---
print(f"📂 Lendo planilha: {ARQUIVO_ENTRADA}")
try:
    df = pd.read_excel(ARQUIVO_ENTRADA)
except FileNotFoundError:
    print(f"❌ ERRO: O arquivo '{ARQUIVO_ENTRADA}' não foi encontrado. Verifique o caminho.")
    exit()

print("🧹 Limpando e corrigindo todos os preços...")
df["Preço (R$)"] = df.get("Preço (R$)", pd.Series(dtype=float)).apply(limpar_preco)
df["Preço Belmicro (R$)"] = df.get("Preço Belmicro (R$)", pd.Series(dtype=float)).apply(limpar_preco)
print("✅ Preços corrigidos e normalizados para float.")

# --- 4. LÓGICA DE SUGESTÃO (CORRIGIDA) ---
print("📊 Gerando relatório de sugestão de preço...")

# --- ETAPA 4.1: CRIAR O "MAPA DE PREÇOS BELMICRO" ---
# (Esta parte continua a mesma, crucial para a lógica)
mapa_precos_belmicro = {}
mapa_avaliacoes_belmicro = {}
for termo, grupo in df.groupby("Termo Pesquisado"):
    belmicro_row = grupo[grupo["Vendedor"].str.contains("belmicro", case=False, na=False)]
    if not belmicro_row.empty:
        mapa_precos_belmicro[termo] = belmicro_row.iloc[0]["Preço Belmicro (R$)"]
        mapa_avaliacoes_belmicro[termo] = belmicro_row.iloc[0]["Avaliação Média"]
print(" -> Mapa de preços de referência da Belmicro foi criado.")

# --- ETAPA 4.2: FILTRAR PRODUTOS APROVADOS (AGORA INCLUINDO BELMICRO) ---
# Em vez de filtrar *apenas* concorrentes, vamos filtrar TODOS os produtos "SIM"
if COLUNA_ANALISE not in df.columns:
    print(f"❌ ERRO: Coluna de filtro '{COLUNA_ANALISE}' não foi encontrada!")
    exit()
    
df_aprovados = df[df[COLUNA_ANALISE].str.startswith('SIM', na=False)].copy()
print(f" -> {len(df_aprovados)} produtos aprovados ('SIM') no total (Belmicro + Concorrentes) serão analisados.")


# --- ETAPA 4.3: GERAR O RELATÓRIO FINAL (LÓGICA ATUALIZADA) ---
linhas = []
for termo, grupo in df_aprovados.groupby("Termo Pesquisado"):
    
    preco_belmicro = mapa_precos_belmicro.get(termo, np.nan)
    avaliacao_belmicro = mapa_avaliacoes_belmicro.get(termo, "-")
    
    # Pega os preços dos concorrentes APROVADOS
    concorrentes_df = grupo[~grupo["Vendedor"].str.contains("belmicro", case=False, na=False)]
    precos_concorrentes = sorted(concorrentes_df["Preço (R$)"].dropna().unique().tolist())
    
    # Lista de todos os preços (Belmicro + Concorrentes APROVADOS)
    precos_todos = list(precos_concorrentes)
    if not np.isnan(preco_belmicro):
        precos_todos.append(preco_belmicro)
    precos_todos = sorted(list(set(precos_todos)))

    if len(precos_todos) == 0:
        continue

    # Posição Belmicro
    posicao = "-"
    if not np.isnan(preco_belmicro) and preco_belmicro in precos_todos:
        posicao = f"{precos_todos.index(preco_belmicro)+1}º de {len(precos_todos)}"

    # Preço sugerido = 3º menor preço
    preco_sugerido = np.nan
    if len(precos_todos) >= 3:
        preco_sugerido = precos_todos[2]
    elif len(precos_todos) == 2:
        preco_sugerido = precos_todos[1]
    else:
        preco_sugerido = precos_todos[0]

    # --- ALTERAÇÃO AQUI: Adiciona a linha da Belmicro primeiro ---
    belmicro_row_df = grupo[grupo["Vendedor"].str.contains("belmicro", case=False, na=False)]
    if not belmicro_row_df.empty:
        belmicro_row = belmicro_row_df.iloc[0] # Pega a primeira linha da Belmicro
        linhas.append({
            "Termo Pesquisado (produto belmicro)": termo,
            "Termo Encontrado": belmicro_row["Nome"],
            "Vendedor Concorrente": "BELMICRO (REFERÊNCIA)", # Identificação clara
            "Preço Concorrente": np.nan, # Deixa o preço concorrente vazio
            "Preço Belmicro Atual": preco_belmicro,
            "Preço Sugerido": preco_sugerido, # Repete o sugerido para o grupo
            "Avaliação Belmicro": avaliacao_belmicro,
            "Avaliação Concorrente": belmicro_row["Avaliação Média"], # Avaliação dela mesma
            "Link da Loja": belmicro_row["Link Loja"],
            "URL": belmicro_row["URL"]
        })
    # --- FIM DA ALTERAÇÃO ---

    # Agora adiciona os concorrentes
    for _, row in concorrentes_df.iterrows():
        linhas.append({
            "Termo Pesquisado (produto belmicro)": termo,
            "Termo Encontrado": row["Nome"],
            "Vendedor Concorrente": row["Vendedor"],
            "Preço Concorrente": row["Preço (R$)"],
            "Preço Belmicro Atual": preco_belmicro,
            "Preço Sugerido": preco_sugerido,
            "Avaliação Belmicro": avaliacao_belmicro,
            "Avaliação Concorrente": row["Avaliação Média"],
            "Link da Loja": row["Link Loja"],
            "URL": row["URL"]
        })
    
    # Linha separadora
    linhas.append({
        "Termo Pesquisado (produto belmicro)": "──────────────────────────────────────────────"
    })

# --- 5. SALVAR RESULTADO ---
print("💾 Salvando planilha final formatada...")
df_final = pd.DataFrame(linhas)

# Renomeia as colunas de preço para a formatação final
df_final = df_final.rename(columns={
    "Preço Concorrente": "Preço Concorrente (R$)",
    "Preço Belmicro Atual": "Preço Belmicro (R$)",
    "Preço Sugerido": "Preço Sugerido (R$)"
})

# Formata as colunas de preço para texto (R$)
# Usamos .applymap para evitar o SettingWithCopyWarning, mas a lógica é a mesma
df_final["Preço Concorrente (R$)"] = df_final["Preço Concorrente (R$)"].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notna(x) else "")
df_final["Preço Belmicro (R$)"] = df_final["Preço Belmicro (R$)"].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notna(x) else "")
df_final["Preço Sugerido (R$)"] = df_final["Preço Sugerido (R$)"].apply(lambda x: f"R$ {x:,.2f} (3º lugar)".replace(",", "X").replace(".", ",").replace("X", ".") if pd.notna(x) else "")

PASTA_SAIDA = os.path.dirname(ARQUIVO_SAIDA)
if not os.path.exists(PASTA_SAIDA):
    os.makedirs(PASTA_SAIDA)
    print(f"✅ Pasta de saída criada em: {PASTA_SAIDA}")

with pd.ExcelWriter(ARQUIVO_SAIDA, engine="openpyxl") as writer:
    df_final.to_excel(writer, index=False, sheet_name="Relatorio_Final")
    ws = writer.sheets["Relatorio_Final"]
    
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid") # Amarelo para Belmicro
    align_center = Alignment(horizontal="center")
    bold_font = Font(bold=True)

    # Cabeçalhos
    for cell in ws["1:1"]:
        cell.font = bold_font
        cell.alignment = align_center

    # Encontra as colunas pelo nome final
    col_letra_sugerido = get_column_letter(df_final.columns.get_loc("Preço Sugerido (R$)") + 1)
    col_letra_conc = get_column_letter(df_final.columns.get_loc("Preço Concorrente (R$)") + 1)
    col_letra_bel = get_column_letter(df_final.columns.get_loc("Preço Belmicro (R$)") + 1)
    col_letra_vendedor = get_column_letter(df_final.columns.get_loc("Vendedor Concorrente") + 1)

    # Loop de formatação
    for row in range(2, ws.max_row + 1):
        # Pinta o Preço Sugerido de verde
        ws[f"{col_letra_sugerido}{row}"].fill = green_fill
        ws[f"{col_letra_sugerido}{row}"].alignment = align_center
        
        # Centraliza os outros preços
        ws[f"{col_letra_conc}{row}"].alignment = align_center
        ws[f"{col_letra_bel}{row}"].alignment = align_center
        
        # --- ALTERAÇÃO: Pinta a linha da Belmicro de Amarelo ---
        cell_vendedor = ws[f"{col_letra_vendedor}{row}"]
        if "BELMICRO (REFERÊNCIA)" in str(cell_vendedor.value):
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col_idx).fill = yellow_fill
        # --- FIM DA ALTERAÇÃO ---
        
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 60)

print(f"✅ Relatório gerado com sucesso: {ARQUIVO_SAIDA}")
